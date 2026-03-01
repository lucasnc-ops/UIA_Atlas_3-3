#!/usr/bin/env python3
"""
excel_to_atlas_sql.py
---------------------
Reads "Resultado para Formatação.xlsx" (141 UIA project submissions) and
generates supabase_import_projects.sql targeting the Atlas 3+3 Supabase schema.

Generated SQL:
  1. Idempotent import user INSERT
  2. Schema extension ALTER TABLE (4 extra columns)
  3. 141 project INSERTs
  4. project_sdgs rows (one per SDG per project, from Col R)

Usage:
    pip install openpyxl
    python scripts/excel_to_atlas_sql.py

Output:
    atlas_33/supabase_import_projects.sql   (ready to paste in Supabase SQL Editor)
"""

import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT  = Path(__file__).parent.parent          # atlas_33/
# Excel lives in the sdg-assessment repo, one level above atlas_33
EXCEL_FILE = REPO_ROOT.parent / "sdg-assessment" / "Resultado para Formatação.xlsx"

# Allow override via env variable (useful in CI)
if os.environ.get("EXCEL_FILE"):
    EXCEL_FILE = Path(os.environ["EXCEL_FILE"])

OUTPUT_SQL = REPO_ROOT / "supabase_import_projects.sql"

IMPORT_EMAIL = "import@uia.archi"

# ---------------------------------------------------------------------------
# Region mapping  (Col H value → uiaregion ENUM)
# ---------------------------------------------------------------------------
REGION_MAP = {
    "region 1": "SECTION_I",
    "region 2": "SECTION_II",
    "region 3": "SECTION_III",
    "region 4": "SECTION_IV",
    "region 5": "SECTION_V",
    "1":        "SECTION_I",
    "2":        "SECTION_II",
    "3":        "SECTION_III",
    "4":        "SECTION_IV",
    "5":        "SECTION_V",
}

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def esc(value) -> str:
    """Escape a value for SQL single-quoted string. Returns NULL for empty."""
    if value is None or str(value).strip() == "":
        return "NULL"
    s = str(value).strip()
    s = s.replace("'", "''")
    return f"'{s}'"


def esc_int(value) -> str:
    """Return an integer literal or NULL."""
    if value is None or str(value).strip() == "":
        return "NULL"
    try:
        return str(int(float(str(value).strip())))
    except (ValueError, TypeError):
        return "NULL"


def esc_float(value) -> str:
    """Return a float literal or NULL."""
    if value is None or str(value).strip() == "":
        return "NULL"
    try:
        return str(float(str(value).strip()))
    except (ValueError, TypeError):
        return "NULL"


def map_region(value) -> str:
    """Map Col H region string to uiaregion ENUM value (quoted), or NULL."""
    if value is None or str(value).strip() == "":
        return "NULL"
    key = str(value).strip().lower()
    mapped = REGION_MAP.get(key)
    if mapped:
        return f"'{mapped}'"
    # Try extracting digit from e.g. "Region 3"
    m = re.search(r"\d", key)
    if m:
        digit = m.group()
        mapped = REGION_MAP.get(digit)
        if mapped:
            return f"'{mapped}'"
    return "NULL"  # unknown — caller will fallback to SECTION_I


def split_city_country(location_val) -> tuple[str, str]:
    """
    Split 'City, Country' on the LAST comma.
    Returns (city_sql, country_sql) as SQL-escaped strings.
    Falls back to (esc(location_val), 'Unknown') if no comma found.
    """
    if location_val is None or str(location_val).strip() == "":
        return ("NULL", "NULL")
    raw = str(location_val).strip()
    if "," in raw:
        idx = raw.rfind(",")
        city    = raw[:idx].strip()
        country = raw[idx + 1:].strip()
        return (esc(city), esc(country))
    # No comma — treat the whole string as city
    return (esc(raw), "'Unknown'")


def parse_sdgs(value) -> list[int]:
    """Parse '1; 3; 7; 10; 13' → [1, 3, 7, 10, 13]."""
    if value is None or str(value).strip() == "":
        return []
    parts = re.split(r"[;,\s]+", str(value).strip())
    sdgs = []
    for p in parts:
        p = p.strip()
        if p.isdigit():
            n = int(p)
            if 1 <= n <= 17:
                sdgs.append(n)
    return sorted(set(sdgs))


def cell(row, col_index):
    """Return cell value by 0-based column index."""
    return row[col_index].value if col_index < len(row) else None


# ---------------------------------------------------------------------------
# Load workbook
# ---------------------------------------------------------------------------

if not EXCEL_FILE.exists():
    print(f"ERROR: Excel file not found:\n  {EXCEL_FILE}")
    print()
    print("Options:")
    print("  1. Copy the file to that path, OR")
    print("  2. Set env variable: EXCEL_FILE=/path/to/file.xlsx")
    sys.exit(1)

print(f"Loading: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Try sheet name variants
sheet_name = None
for candidate in ["Planilha1", "Sheet1", "Projects", wb.sheetnames[0]]:
    if candidate in wb.sheetnames:
        sheet_name = candidate
        break

ws = wb[sheet_name]
rows = list(ws.iter_rows(min_row=2))  # skip header row
print(f"  Sheet: '{sheet_name}' — rows (excluding header): {len(rows)}")

# ---------------------------------------------------------------------------
# Build SQL
# ---------------------------------------------------------------------------

lines = []

lines += [
    "-- =============================================================================",
    "-- supabase_import_projects.sql",
    "-- Generated by scripts/excel_to_atlas_sql.py",
    "-- Imports 141 UIA project submissions (Atlas 3+3 schema)",
    "-- Source: Resultado para Formatação.xlsx",
    "-- =============================================================================",
    "",
    "-- Run this entire script in the Supabase SQL Editor.",
    "-- It is idempotent for the user INSERT; project rows are appended each run.",
    "-- To start fresh: truncate projects CASCADE; first.",
    "",
]

# ---------------------------------------------------------------------------
# Step 1 — Schema extensions (idempotent)
# ---------------------------------------------------------------------------
lines += [
    "-- ---------------------------------------------------------------------------",
    "-- Step 1: Schema extensions (idempotent)",
    "-- ---------------------------------------------------------------------------",
    "",
    "ALTER TABLE projects",
    "    ADD COLUMN IF NOT EXISTS external_code  VARCHAR(16),",
    "    ADD COLUMN IF NOT EXISTS authors        TEXT,",
    "    ADD COLUMN IF NOT EXISTS size_sqm       FLOAT,",
    "    ADD COLUMN IF NOT EXISTS year_completed INTEGER;",
    "",
    "CREATE INDEX IF NOT EXISTS ix_projects_external_code ON projects(external_code);",
    "",
]

# ---------------------------------------------------------------------------
# Step 2 — Import user (idempotent)
# ---------------------------------------------------------------------------
lines += [
    "-- ---------------------------------------------------------------------------",
    "-- Step 2: Import user (idempotent)",
    "-- ---------------------------------------------------------------------------",
    "",
    "INSERT INTO users (email, hashed_password, role)",
    f"VALUES ({esc(IMPORT_EMAIL)}, 'not-a-real-hash', 'ADMIN')",
    "ON CONFLICT (email) DO NOTHING;",
    "",
]

# ---------------------------------------------------------------------------
# Step 3 — Insert 141 projects
# ---------------------------------------------------------------------------
lines += [
    "-- ---------------------------------------------------------------------------",
    "-- Step 3: Insert projects",
    "-- ---------------------------------------------------------------------------",
    "",
]

project_info = []   # [(external_code, sdg_list), ...]
skipped = 0

for i, row in enumerate(rows, start=1):
    # ---- Read cells (0-based column index) ---------------------------------
    selected_flag = cell(row, 0)    # A: "X" if selected
    # B (index 1): primary SDG — not used; Col R has full list
    code          = cell(row, 2)    # C: external code (EFP1, IFF35, …)
    # D (index 3): sequential int — skip
    name          = cell(row, 4)    # E: project name
    authors_val   = cell(row, 5)    # F: architect names
    org_val       = cell(row, 6)    # G: organization
    region_val    = cell(row, 7)    # H: "Region 1" … "Region 5"
    # I (index 8): client — skip
    location_val  = cell(row, 9)    # J: "City, Country"
    # K (index 10): geographic ref — skip
    size_val      = cell(row, 11)   # L: built area m²
    year_val      = cell(row, 12)   # M: year of conclusion
    # N (index 13): empty — skip
    brief_val     = cell(row, 14)   # O: 200-char description
    detail_val    = cell(row, 15)   # P: 500-char summary
    success_val   = cell(row, 16)   # Q: SDG connection explanation
    sdg_list_val  = cell(row, 17)   # R: SDG list "1; 3; 7; …"
    # S (index 18): skip
    contact_name  = cell(row, 19)   # T: applicant name
    # U, V, W: skip
    contact_email = cell(row, 23)   # X: business email

    # Skip completely empty rows
    if name is None and code is None:
        skipped += 1
        continue

    # ---- Derived values ----------------------------------------------------
    ext_code = str(code).strip() if code else f"ROW{i}"

    # workflow_status: "X" in Col A → APPROVED, else SUBMITTED
    if selected_flag is not None and str(selected_flag).strip().upper() == "X":
        wf_status = "'APPROVED'"
    else:
        wf_status = "'SUBMITTED'"

    # uia_region: map or fallback
    region_sql = map_region(region_val)
    if region_sql == "NULL":
        region_sql = "'SECTION_I'"   # safe fallback; unknown region

    # city / country from Col J
    city_sql, country_sql = split_city_country(location_val)
    if city_sql == "NULL":
        city_sql = "'Unknown'"
    if country_sql == "NULL":
        country_sql = "'Unknown'"

    # contact_person fallback
    contact_person_sql = esc(contact_name) if contact_name else "'Unknown'"

    # contact_email fallback (must be something)
    contact_email_sql = esc(contact_email) if contact_email else f"'noemail+{ext_code.lower()}@uia.archi'"

    # brief_description: required NOT NULL
    brief_sql = esc(brief_val) if brief_val and str(brief_val).strip() else esc(name)

    # detailed_description: required NOT NULL
    detail_sql = esc(detail_val) if detail_val and str(detail_val).strip() else brief_sql

    # success_factors: required NOT NULL
    success_sql = esc(success_val) if success_val and str(success_val).strip() else "'No data'"

    # organization_name: required NOT NULL
    org_sql = esc(org_val) if org_val and str(org_val).strip() else "'Unknown'"

    # project_name: required NOT NULL
    name_sql = esc(name) if name and str(name).strip() else esc(ext_code)

    sdgs = parse_sdgs(sdg_list_val)
    project_info.append((ext_code, sdgs))

    lines.append(f"-- Row {i}: {ext_code}")
    lines += [
        "INSERT INTO projects (",
        "    organization_name, contact_person, contact_email, project_name,",
        "    project_status, workflow_status,",
        "    funding_needed, funding_spent,",
        "    uia_region, city, country,",
        "    latitude, longitude,",
        "    brief_description, detailed_description, success_factors,",
        "    gdpr_consent,",
        "    external_code, authors, size_sqm, year_completed",
        ") VALUES (",
        f"    {org_sql}, {contact_person_sql}, {contact_email_sql}, {name_sql},",
        f"    'IMPLEMENTED', {wf_status},",
        f"    0.0, 0.0,",
        f"    {region_sql}, {city_sql}, {country_sql},",
        f"    NULL, NULL,",
        f"    {brief_sql}, {detail_sql}, {success_sql},",
        f"    TRUE,",
        f"    {esc(ext_code)}, {esc(authors_val)}, {esc_float(size_val)}, {esc_int(year_val)}",
        ");",
        "",
    ]

# ---------------------------------------------------------------------------
# Step 4 — project_sdgs
# ---------------------------------------------------------------------------
lines += [
    "-- ---------------------------------------------------------------------------",
    "-- Step 4: project_sdgs (one row per SDG per project)",
    "-- ---------------------------------------------------------------------------",
    "",
]

for ext_code, sdgs in project_info:
    if not sdgs:
        continue
    for sdg_num in sdgs:
        lines += [
            "INSERT INTO project_sdgs (project_id, sdg_number)",
            f"SELECT id, {sdg_num} FROM projects WHERE external_code = {esc(ext_code)}",
            "ON CONFLICT DO NOTHING;",
        ]
    lines.append("")

# ---------------------------------------------------------------------------
# Verification queries
# ---------------------------------------------------------------------------
lines += [
    "-- ---------------------------------------------------------------------------",
    "-- Verification (run after import):",
    "-- ---------------------------------------------------------------------------",
    "-- SELECT COUNT(*) FROM projects;       -- expect 141",
    "-- SELECT COUNT(*) FROM project_sdgs;   -- expect ~697",
    "-- SELECT workflow_status, COUNT(*) FROM projects GROUP BY workflow_status;",
    "-- -- expect: APPROVED=96, SUBMITTED=45",
    "-- SELECT project_name, external_code, city, country",
    "--   FROM projects WHERE external_code = 'EFP1';",
    "",
]

# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------
output = "\n".join(lines)
OUTPUT_SQL.write_text(output, encoding="utf-8")

total_sdg_pairs = sum(len(sdgs) for _, sdgs in project_info)
approved = sum(1 for row in rows if cell(row, 4) is not None or cell(row, 2) is not None
               if cell(row, 0) is not None and str(cell(row, 0)).strip().upper() == "X")

print(f"\nGenerated: {OUTPUT_SQL}")
print(f"  Projects inserted : {len(project_info)}")
print(f"  Skipped (empty)   : {skipped}")
print(f"  project_sdgs rows : {total_sdg_pairs}")
print(f"  Output file size  : {OUTPUT_SQL.stat().st_size:,} bytes")
print()
print("Next step: paste supabase_import_projects.sql into the Supabase SQL Editor.")

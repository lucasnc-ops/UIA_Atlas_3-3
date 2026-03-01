#!/usr/bin/env python3
"""
generate_coord_updates.py
--------------------------
Reads "Resultado para Formatação.xlsx" (141 UIA project submissions),
maps each project's raw city/country to user-verified geocoordinates,
and generates supabase_update_coordinates.sql.

Generated SQL updates, keyed on external_code (unique index):
    - latitude, longitude
    - city, country  (cleaned / standardised values)

Skips IFF6 ("Everywhere") and IFF29 (blank location) — leaves them NULL.

Usage:
    pip install openpyxl
    python scripts/generate_coord_updates.py

Output:
    atlas_33/supabase_update_coordinates.sql
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT  = Path(__file__).parent.parent          # atlas_33/
EXCEL_FILE = REPO_ROOT.parent / "sdg-assessment" / "Resultado para Formatação.xlsx"

if os.environ.get("EXCEL_FILE"):
    EXCEL_FILE = Path(os.environ["EXCEL_FILE"])

OUTPUT_SQL = REPO_ROOT / "supabase_update_coordinates.sql"

# ---------------------------------------------------------------------------
# User-verified coordinate lookup
#
# Keys   : (city_raw, country_raw) — exact strings produced by the parser
#           (raw Excel cell split on last ASCII comma, stripped)
# Values : (latitude, longitude, clean_city, clean_country)
#           OR None  →  skip this project (no geocodable location)
# ---------------------------------------------------------------------------
COORD_MAP = {
    # (city_raw, country_raw): (lat, lon, clean_city, clean_country)

    # Blank / no-location sentinel (IFF29)
    ("", ""):                                               None,

    # --- Austria ---
    ("Murstetten",                      "Austria"):         ( 48.233300,   15.816700, "Murstetten",           "Austria"),
    ("Pressbaum/Vienna",                "Austria"):         ( 48.182288,   16.075958, "Vienna",               "Austria"),

    # --- Bangladesh ---
    ("Dhaka",                           "Bangladesh"):      ( 23.766169,   90.390771, "Dhaka",                "Bangladesh"),
    ("Gopalganj",                       "Bangladesh"):      ( 23.013014,   89.822405, "Gopalganj",            "Bangladesh"),
    ("Jhenaidah",                       "Bangladesh"):      ( 23.633841,   89.066856, "Jhenaidah",            "Bangladesh"),
    ("Kuakata",                         "Bangladesh"):      ( 21.815347,   90.140874, "Kuakata",              "Bangladesh"),

    # --- Brazil (raw country = "Brazil") ---
    ("Canoas, Rio Grande do Sul",       "Brazil"):          (-29.920059,  -51.179004, "Canoas",               "Brazil"),
    ("Foz do Iguaçu",                   "Brazil"):          (-25.538124,  -54.585437, "Foz do Iguaçu",        "Brazil"),
    ("Juriti",                          "Brazil"):          ( -2.152800,  -56.090733, "Juriti",               "Brazil"),
    ("Manhumirim",                      "Brazil"):          (-20.358367,  -41.958920, "Manhumirim",           "Brazil"),
    ("Natal",                           "Brazil"):          ( -5.800832,  -35.225822, "Natal",                "Brazil"),
    ("Rio de Janeiro and Niteroi",      "Brazil"):          (-22.902094,  -43.128199, "Rio de Janeiro",       "Brazil"),
    ("Sobral",                          "Brazil"):          ( -3.670293,  -40.314005, "Sobral",               "Brazil"),
    ("São Luís",                        "Brazil"):          ( -2.529973,  -44.300193, "São Luís",             "Brazil"),
    ("São Paulo",                       "Brazil"):          (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    # raw country = "Brasil"
    ("Nazaré Paulista",                 "Brasil"):          (-23.183331,  -46.399813, "Nazaré Paulista",      "Brazil"),
    ("Sao Paulo",                       "Brasil"):          (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    ("São Paulo",                       "Brasil"):          (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    ("Valinhos",                        "Brasil"):          (-22.970559,  -46.995831, "Valinhos",             "Brazil"),
    # raw country = "Brazil." (trailing period)
    ("São Paulo",                       "Brazil."):         (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    # raw country = "SP"
    ("Sao Paulo",                       "SP"):              (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    # raw country = "Maranhão - Brazil"
    ("São Luís",                        "Maranhão - Brazil"):( -2.529973,  -44.300193, "São Luís",            "Brazil"),

    # --- Canada ---
    ("Calgary",                         "Canada"):          ( 51.048735, -114.072474, "Calgary",              "Canada"),
    ("Town of Athabasca",               "Canada"):          ( 54.716497, -113.284176, "Athabasca",            "Canada"),

    # --- Cape Verde ---
    ("Brava and Fogo Islands",          "Cape Verde"):      ( 14.896567,  -24.541500, "Brava & Fogo Islands", "Cape Verde"),
    ("Sal Island",                      "Cape Verde"):      ( 16.775533,  -22.921133, "Sal Island",           "Cape Verde"),
    ("São Nicolau, São Vicente and Santi Antão",
                                        "Cape Verde"):      ( 16.835000,  -24.807600, "São Nicolau",          "Cape Verde"),

    # --- Chile ---
    ("Concepción, Region del Bio-Bio",  "Chile"):           (-36.827393,  -73.050370, "Concepción",           "Chile"),

    # --- China ---
    ("Baoxi town",                      "China"):           ( 28.006110,  118.760280, "Baoxi",                "China"),
    ("Beijing",                         "China"):           ( 39.912433,  116.401033, "Beijing",              "China"),
    ("Changyuan, Hunan Province",       "China"):           ( 35.212500,  114.735000, "Changyuan",            "China"),
    ("Chengmai, Hainan Province",       "China"):           ( 19.738500,  110.006767, "Chengmai",             "China"),
    ("Chongli",                         "China"):           ( 40.998857,  115.261397, "Chongli",              "China"),
    ("Dali",                            "China"):           ( 25.606500,  100.267600, "Dali",                 "China"),
    ("Fengning",                        "China"):           ( 41.200000,  116.650000, "Fengning",             "China"),
    ("Fujian Province",                 "China"):           ( 26.484000,  117.925000, "Fujian",               "China"),
    ("Gaobeidian City, Hebei Province", "China"):           ( 39.284860,  115.954605, "Gaobeidian",           "China"),
    ("Guian New Area, Guizhou province","China"):           ( 26.468800,  106.483100, "Guian New Area",       "China"),
    ("Nanjing",                         "China"):           ( 32.053418,  118.776849, "Nanjing",              "China"),
    ("Shangcun Village, Xuancheng city","China"):           ( 29.945833,  118.748056, "Shangcun Village",     "China"),
    ("Shanghai",                        "China"):           ( 31.222934,  121.461763, "Shanghai",             "China"),
    ("Wuzhen",                          "China"):           ( 30.742453,  120.488833, "Wuzhen",               "China"),
    ("Xicheng District, Beijing",       "China"):           ( 39.910309,  116.365435, "Beijing",              "China"),
    ("Yanqing County, Beijing",         "China"):           ( 40.420000,  116.083336, "Yanqing",              "China"),
    ("Yixing, Jiangsu Province",        "China"):           ( 31.360000,  119.815000, "Yixing",               "China"),
    ("Yueyang County, Hunan Province",  "China"):           ( 29.144600,  113.116100, "Yueyang",              "China"),
    ("Zhuhai",                          "China"):           ( 22.274046,  113.572230, "Zhuhai",               "China"),
    # raw country = "P.R.China"
    ("Nanjing, Jiangsu province",       "P.R.China"):       ( 32.053418,  118.776849, "Nanjing",              "China"),

    # --- Colombia (raw country = "Colômbia") ---
    ("Bogotá",                          "Colômbia"):        (  4.673345,  -74.073339, "Bogotá",               "Colombia"),

    # --- Denmark ---
    ("Copenhagen",                      "Denmark"):         ( 55.679521,   12.567034, "Copenhagen",           "Denmark"),

    # --- Egypt ---
    ("Fayoum",                          "Egypt"):           ( 29.311889,   30.826800, "Fayoum",               "Egypt"),
    # city/country swapped in source: city="Egypt", country="Cairo"
    ("Egypt",                           "Cairo"):           ( 30.046784,   31.239582, "Cairo",                "Egypt"),
    # raw country = "Egypt." (trailing period)
    ("Cairo",                           "Egypt."):          ( 30.046784,   31.239582, "Cairo",                "Egypt"),
    # multi-city; raw country = "New Cairo"
    ("Egypt cities; Aga, New Atfeh, New Obour",
                                        "New Cairo"):       ( 30.030885,   31.421844, "New Cairo",            "Egypt"),

    # --- Germany ---
    ("Munich",                          "Germany"):         ( 48.141560,   11.567313, "Munich",               "Germany"),

    # --- Greece ---
    ("Heraklion, Crete",                "Greece"):          ( 35.339887,   25.141917, "Heraklion",            "Greece"),
    ("Thessaloniki",                    "Greece"):          ( 40.637410,   22.940960, "Thessaloniki",         "Greece"),

    # --- India ---
    ("Alibaug",                         "India"):           ( 18.640548,   72.872215, "Alibaug",              "India"),
    ("Bengaluru",                       "India"):           ( 12.974220,   77.593185, "Bengaluru",            "India"),
    ("Bhubaneswar",                     "India"):           ( 20.296059,   85.824539, "Bhubaneswar",          "India"),
    ("Global application, developed in Bengaluru",
                                        "India"):           ( 12.974220,   77.593185, "Bengaluru",            "India"),
    ("New Delhi",                       "India"):           ( 28.635773,   77.224433, "New Delhi",            "India"),
    ("Pune",                            "India"):           ( 18.518909,   73.856116, "Pune",                 "India"),
    # raw country = "INDIA" (all caps)
    ("LEH",                             "INDIA"):           ( 34.160889,   77.581696, "Leh",                  "India"),

    # --- Italy ---
    ("Bisceglie",                       "Italy"):           ( 41.229019,   16.495083, "Bisceglie",            "Italy"),
    ("Maida",                           "Italy"):           ( 38.859613,   16.363700, "Maida",                "Italy"),
    ("Milan",                           "Italy"):           ( 45.464654,    9.188315, "Milan",                "Italy"),
    ("Turin",                           "Italy"):           ( 45.085719,    7.705418, "Turin",                "Italy"),

    # --- Japan ---
    ("Kusatsu",                         "Japan"):           ( 35.017480,  135.955594, "Kusatsu",              "Japan"),

    # --- Malaysia ---
    ("Ipoh",                            "Malaysia"):        (  4.593420,  101.087676, "Ipoh",                 "Malaysia"),

    # --- Mexico (raw country = "México") ---
    ("Guadalajara",                     "México"):          ( 20.662066, -103.363739, "Guadalajara",          "Mexico"),

    # --- Mongolia ---
    ("Ulaanbaatar",                     "Mongolia"):        ( 47.898009,  106.910015, "Ulaanbaatar",          "Mongolia"),

    # --- Morocco ---
    ("Benguerir",                       "Morocco"):         ( 32.240335,   -7.953709, "Benguerir",            "Morocco"),

    # --- North Macedonia (raw country = "N. Macedonia") ---
    ("Prilep",                          "N. Macedonia"):    ( 41.347000,   21.556950, "Prilep",               "North Macedonia"),

    # --- Pakistan ---
    ("Jacobabad (Sindh)",               "Pakistan"):        ( 28.279300,   68.437270, "Jacobabad",            "Pakistan"),
    ("Karachi",                         "Pakistan"):        ( 24.891140,   67.046580, "Karachi",              "Pakistan"),
    ("Khurrianwala, Punjab",            "Pakistan"):        ( 31.517220,   73.266670, "Khurrianwala",         "Pakistan"),
    ("Uthal (Balochistan)",             "Pakistan"):        ( 25.803900,   66.619490, "Uthal",                "Pakistan"),
    # raw country = "PAKISTAN" (all caps)
    ("KARACHI",                         "PAKISTAN"):        ( 24.891140,   67.046580, "Karachi",              "Pakistan"),

    # --- Peru ---
    ("Arequipa",                        "Peru"):            (-16.397000,  -71.537480, "Arequipa",             "Peru"),

    # --- Romania ---
    ("Bucharest",                       "Romania"):         ( 44.432680,   26.102010, "Bucharest",            "Romania"),
    ("Plosca village, Teliucu, Hunedoara county",
                                        "Romania"):         ( 44.033000,   25.117000, "Plosca",               "Romania"),
    ("Reci, Covasna County",            "Romania"):         ( 45.850000,   25.933000, "Reci",                 "Romania"),

    # --- South Africa ---
    ("Bavianskloof Valley",             "South Africa"):    (-33.680270,   24.514810, "Bavianskloof Valley",  "South Africa"),

    # --- Spain ---
    ("Burgos",                          "Spain"):           ( 42.341060,   -3.701840, "Burgos",               "Spain"),
    ("Logroño",                         "Spain"):           ( 42.464110,   -2.448520, "Logroño",              "Spain"),
    ("Pamplona",                        "Spain"):           ( 42.816870,   -1.643230, "Pamplona",             "Spain"),
    ("Valladolid",                      "Spain"):           ( 41.653510,   -4.723780, "Valladolid",           "Spain"),
    # raw country = "Navarra"
    ("Pamplona",                        "Navarra"):         ( 42.816870,   -1.643230, "Pamplona",             "Spain"),
    # raw country = "Sapin" (typo)
    ("Burgos",                          "Sapin"):           ( 42.341060,   -3.701840, "Burgos",               "Spain"),

    # --- Uganda ---
    ("Karamoja",                        "Uganda"):          (  2.750000,   34.250000, "Karamoja",             "Uganda"),
    ("Wakiso",                          "Uganda"):          (  0.404440,   32.459440, "Wakiso",               "Uganda"),

    # --- United States ---
    ("New York City",                   "United States of America"):
                                                            ( 40.714270,  -74.005970, "New York City",        "United States of America"),

    # --- Uruguay ---
    ("Montevideo",                      "Uruguay"):         (-34.894443,  -56.173933, "Montevideo",           "Uruguay"),

    # --- Vietnam ---
    ("Hanoi",                           "Vietnam"):         ( 21.031477,  105.844719, "Hanoi",                "Vietnam"),
    ("Ha Noi",                          "Viet Nam"):        ( 21.031477,  105.844719, "Hanoi",                "Vietnam"),

    # -------------------------------------------------------------------------
    # "Unknown" country entries
    # These rows had no comma in the Excel cell, so split_city_country()
    # returned (whole_string, "Unknown").
    # -------------------------------------------------------------------------
    ("Belo Horizonte-MG",               "Unknown"):         (-19.918219,  -43.938831, "Belo Horizonte",       "Brazil"),
    ("Bengaluru and India",             "Unknown"):         ( 12.974220,   77.593185, "Bengaluru",            "India"),
    ("Boa Vista/Brazil",                "Unknown"):         (  2.820337,  -60.671799, "Boa Vista",            "Brazil"),
    # Fullwidth comma U+FF0C — split_city_country sees no ASCII comma
    ("Chengdu\uff0cChina",              "Unknown"):         ( 30.659900,  104.063400, "Chengdu",              "China"),
    ("Dubaï / United Arab States",      "Unknown"):         ( 25.270678,   55.299413, "Dubai",                "United Arab Emirates"),
    ("Everywhere",                      "Unknown"):         None,   # IFF6 — no geocodable location
    ("Hong Kong",                       "Unknown"):         ( 22.296807,  114.161627, "Hong Kong",            "China"),
    ("Jordan Irbid city",               "Unknown"):         ( 32.554655,   35.849792, "Irbid",                "Jordan"),
    ("Kunming\uff0cChina",              "Unknown"):         ( 25.039260,  102.718887, "Kunming",              "China"),
    ("Milan Italy",                     "Unknown"):         ( 45.464654,    9.188315, "Milan",                "Italy"),
    ("Pisa (PI) - Italia",              "Unknown"):         ( 43.718811,   10.398847, "Pisa",                 "Italy"),
    ("Porto Alegre/ Brazil",            "Unknown"):         (-30.029726,  -51.229265, "Porto Alegre",         "Brazil"),
    ("Pune & India",                    "Unknown"):         ( 18.518909,   73.856116, "Pune",                 "India"),
    ("Southwest rural China",           "Unknown"):         ( 26.000000,  104.000000, "Southwest China",      "China"),
    ("São Paulo - Brazil",              "Unknown"):         (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    ("São Paulo / Brazil",              "Unknown"):         (-23.543931,  -46.631567, "São Paulo",            "Brazil"),
    ("Uthal - Pakistan",                "Unknown"):         ( 25.804807,   66.620280, "Uthal",                "Pakistan"),
    ("Valladolid. Spain",               "Unknown"):         ( 41.653293,   -4.724193, "Valladolid",           "Spain"),
    ("Zaghouan / Tunisia",              "Unknown"):         ( 36.404973,   10.142706, "Zaghouan",             "Tunisia"),
    ("karachi and pakistan",            "Unknown"):         ( 24.891000,   67.046500, "Karachi",              "Pakistan"),
    ("on the border between São Paulo and Diadema / Brazil",
                                        "Unknown"):         (-23.686500,  -46.623400, "São Paulo/Diadema",    "Brazil"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sql_str(s: str) -> str:
    """Return SQL single-quoted string with escaped single quotes."""
    return "'" + s.replace("'", "''") + "'"


def get_raw_city_country(location_val) -> tuple[str, str]:
    """
    Mirror the logic of split_city_country() in excel_to_atlas_sql.py,
    but return raw unescaped strings for dict lookup.
    Splits on the LAST ASCII comma.
    """
    if location_val is None or str(location_val).strip() == "":
        return ("", "")
    raw = str(location_val).strip()
    if "," in raw:          # ASCII comma U+002C only
        idx = raw.rfind(",")
        city    = raw[:idx].strip()
        country = raw[idx + 1:].strip()
        return (city, country)
    # No comma — whole string is the city, country unknown
    return (raw, "Unknown")


def cell(row, col_index):
    """Return cell value by 0-based column index."""
    return row[col_index].value if col_index < len(row) else None


# ---------------------------------------------------------------------------
# Load workbook
# ---------------------------------------------------------------------------

if not EXCEL_FILE.exists():
    print(f"ERROR: Excel file not found:\n  {EXCEL_FILE}")
    print("Set env variable EXCEL_FILE=/path/to/file.xlsx to override.")
    sys.exit(1)

print(f"Loading: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

sheet_name = None
for candidate in ["Planilha1", "Sheet1", "Projects", wb.sheetnames[0]]:
    if candidate in wb.sheetnames:
        sheet_name = candidate
        break

ws = wb[sheet_name]
rows = list(ws.iter_rows(min_row=2))  # skip header
print(f"  Sheet: '{sheet_name}' — rows (excluding header): {len(rows)}")

# ---------------------------------------------------------------------------
# Generate UPDATE statements
# ---------------------------------------------------------------------------

lines = [
    "-- =============================================================================",
    "-- supabase_update_coordinates.sql",
    "-- Generated by scripts/generate_coord_updates.py",
    "-- Sets latitude, longitude, city (clean), country (clean) for all 141 projects",
    "-- Source coordinates: user-verified (116 unique city/country pairs)",
    "-- =============================================================================",
    "",
    "-- Run in the Supabase SQL Editor after supabase_import_projects.sql.",
    "-- Idempotent — safe to re-run.",
    "",
]

updated  = 0
skipped  = 0
missing  = 0
warnings = []

for i, row in enumerate(rows, start=1):
    code         = cell(row, 2)    # Col C: external_code
    name         = cell(row, 4)    # Col E: project name
    location_val = cell(row, 9)    # Col J: "City, Country"

    # Skip completely empty rows
    if name is None and code is None:
        continue

    ext_code = str(code).strip() if code else f"ROW{i}"
    city_raw, country_raw = get_raw_city_country(location_val)
    key = (city_raw, country_raw)

    if key not in COORD_MAP:
        msg = f"  -- WARNING row {i} ({ext_code}): no mapping for {key!r}"
        warnings.append(msg)
        lines.append(msg)
        missing += 1
        continue

    coords = COORD_MAP[key]

    if coords is None:
        # Explicitly no location (IFF6 "Everywhere", IFF29 blank)
        lines.append(f"-- SKIP {ext_code}: no geocodable location ({city_raw!r}, {country_raw!r})")
        skipped += 1
        continue

    lat, lon, clean_city, clean_country = coords

    lines += [
        f"-- Row {i}: {ext_code}  ({clean_city}, {clean_country})",
        f"UPDATE projects",
        f"    SET latitude  = {lat},",
        f"        longitude = {lon},",
        f"        city      = {sql_str(clean_city)},",
        f"        country   = {sql_str(clean_country)}",
        f"    WHERE external_code = {sql_str(ext_code)};",
        "",
    ]
    updated += 1

# ---------------------------------------------------------------------------
# Verification queries
# ---------------------------------------------------------------------------
lines += [
    "-- =============================================================================",
    "-- Verification (run after applying this script):",
    "-- =============================================================================",
    f"-- Total UPDATEs generated : {updated}",
    f"-- Skipped (no location)   : {skipped}",
    f"-- Missing from map        : {missing}",
    "--",
    "-- SELECT COUNT(*) FROM projects WHERE latitude IS NOT NULL;   -- expect ~139",
    "-- SELECT COUNT(*) FROM projects WHERE latitude IS NULL;        -- expect ~2",
    "-- SELECT external_code, city, country, latitude, longitude",
    "--   FROM projects WHERE external_code IN ('EFP1','LDP14','LDP67','IFF53');",
    "-- SELECT country, COUNT(*) FROM projects",
    "--   WHERE external_code IS NOT NULL",
    "--   GROUP BY country ORDER BY COUNT(*) DESC;",
    "",
]

# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------
output = "\n".join(lines)
OUTPUT_SQL.write_text(output, encoding="utf-8")

print(f"\nGenerated: {OUTPUT_SQL}")
print(f"  UPDATE statements : {updated}")
print(f"  Skipped (no loc)  : {skipped}")
print(f"  Missing from map  : {missing}")
if warnings:
    print("\nWarnings:")
    for w in warnings:
        print(w)
print("\nNext step: paste supabase_update_coordinates.sql into the Supabase SQL Editor.")
